from openpyxl import load_workbook
import pandas as pd
def standard (list1):
    sumOfValue = 0
    for listItem in list1:
        sumOfValue += float(listItem[1])
        #print(float(listItem[1]))
    averageValue = sumOfValue/len(list1)
    #print(averageValue)
    standard = averageValue/6
    #print(standard)
    standardList = []
    count = 0
    value = 0
    for listItem in list1:
        value = int(float(listItem[1])/standard)
        if value > 5:
            value = 5
        standardList.append((listItem[0], value))
        #print(standardList[count])
        count += 1
    # print(standardList)

    return(standardList)
def standard2 (list1):
    sumOfValue = 0
    for listItem in list1:
        sumOfValue += float(listItem[1])
        #print(float(listItem[1]))
    averageValue = sumOfValue/len(list1)
    #print(averageValue)
    standard = averageValue/2.5
    #print(standard)
    standardList = []
    count = 0
    value = 0
    for listItem in list1:
        value = int(float(listItem[1])/standard)
        if value > 5:
            value = 5
        standardList.append((listItem[0], value))
        #print(standardList[count])
        count += 1
    # print(standardList)

    return(standardList)

wb = load_workbook('DSS資料v3.xlsx')
a_sheet = wb.get_sheet_by_name('中壢')
b_sheet = wb.get_sheet_by_name('桃園')
qq = list()
prop = "四大超商"
for j in range(2, 8):
    if prop == a_sheet.cell(row=3, column=j).value:
        for i in range(6, 91):
            qq.append((a_sheet.cell(row=i, column=1).value, a_sheet.cell(row=i, column=j).value))
        for k in range(6, 82):
            qq.append((b_sheet.cell(row=k, column=1).value, b_sheet.cell(row=k, column=j).value))
print(qq)
print(standard(qq))
print(standard2(qq))