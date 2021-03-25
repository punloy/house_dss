import pyodbc 
import string
import pandas as pd
from flask import Flask, render_template, url_for, request
from openpyxl import load_workbook

#conn = pyodbc.connect(r'Driver={SQL Server};Server=DESKTOP-4Q06A90;Database=house;Trusted_Connection=yes;')

#標準化
def standard (list1):
    sumOfValue = 0
    for listItem in list1:
        sumOfValue += float(listItem[1])
        #print(float(listItem[1]))
    averageValue = sumOfValue/len(list1)
    #print(averageValue)
    standard = averageValue/6
    #print(standard)
    standardList = []
    count = 0
    value = 0
    for listItem in list1:
        value = int(float(listItem[1])/standard)
        if value > 5:
            value = 5
        standardList.append((listItem[0], value))
        #print(standardList[count])
        count += 1
    # print(standardList)

    return(standardList)

#傳入權重及user選擇的屬性
# 生活機能
def callLife(scale, prop):

    """query = "select neighborhood, " + prop +" from house.dbo.life;"
    cursor = conn.cursor()
    cursor.execute(query)"""
    print("a")
    stand = []
    wb = load_workbook('DSS資料v3.xlsx')
    a_sheet = wb.get_sheet_by_name('中壢')
    b_sheet = wb.get_sheet_by_name('桃園')
    print(a_sheet.cell(row=3, column=2).value)
    print(prop)
    for j in range(2, 8):
        if prop == a_sheet.cell(row=4, column=j).value:
            print("b")
            for i in range(6, 91):
                print(a_sheet.cell(row=i, column=j).value)
                stand.append((a_sheet.cell(row=i, column=1).value, a_sheet.cell(row=i, column=j).value))
            for k in range(6, 82):
                stand.append((b_sheet.cell(row=k, column=1).value, b_sheet.cell(row=k, column=j).value))
    print(stand)
    life = []

    """for row in cursor:
        stand.append(row)"""
        
    stand = standard(stand)
    
    for x, y in stand:
        y *= scale
        life.append([x, y])
    
    return(life)


# 鄰避設施
def callNeighborSayNo(scale, prop):

    """query = "select neighborhood, " + prop +" from house.dbo.neighborSayNo;"
    cursor = conn.cursor()
    cursor.execute(query)"""
    stand = []
    wb = load_workbook('DSS資料v3.xlsx')
    a_sheet = wb.get_sheet_by_name('中壢')
    b_sheet = wb.get_sheet_by_name('桃園')
    
    for j in range(8, 12):
        if prop == a_sheet.cell(row=4, column=j).value:
            for i in range(6, 91):
                print(a_sheet.cell(row=i, column=j).value)
                stand.append((a_sheet.cell(row=i, column=1).value, a_sheet.cell(row=i, column=j).value))
            for k in range(6, 82):
                stand.append((b_sheet.cell(row=k, column=1).value, b_sheet.cell(row=k, column=j).value))


    SayNo = []
    """for row in cursor:
        stand.append(row)"""
        
    stand = standard(stand)
    
    for x, y in stand:
        y *= scale
        SayNo.append([x, y])

    return(SayNo)

# 公共安全
def callSafety(scale, prop):

    """query = "select neighborhood, " + prop +" from house.dbo.safety;"
    cursor = conn.cursor()
    cursor.execute(query)"""
    stand = list()
    wb = load_workbook('DSS資料v3.xlsx')
    a_sheet = wb.get_sheet_by_name('中壢')
    b_sheet = wb.get_sheet_by_name('桃園')
    for j in range(12, 14):
        if prop == a_sheet.cell(row=4, column=j).value:
            for i in range(6, 91):
                print(a_sheet.cell(row=i, column=j).value)
                stand.append((a_sheet.cell(row=i, column=1).value, a_sheet.cell(row=i, column=j).value))
            for k in range(6, 82):
                stand.append((b_sheet.cell(row=k, column=1).value, b_sheet.cell(row=k, column=j).value))
 
    safety = []
    """for row in cursor:
        stand.append(row)"""
        
    stand = standard(stand)
    
    for x, y in stand:
        y *= scale
        safety.append([x, y])

    return(safety)

# 工資
def callWage(scale, prop):

    """query = "select neighborhood, " + prop +" from house.dbo.wage;"
    cursor = conn.cursor()
    cursor.execute(query)"""
    stand = []
    wb = load_workbook('DSS資料v3.xlsx')
    a_sheet = wb.get_sheet_by_name('中壢')
    b_sheet = wb.get_sheet_by_name('桃園')
    for j in range(14, 16):
        if prop == a_sheet.cell(row=4, column=j).value:
            for i in range(6, 91):
                print(a_sheet.cell(row=i, column=j).value)
                stand.append((a_sheet.cell(row=i, column=1).value, a_sheet.cell(row=i, column=j).value))
            for k in range(6, 82):
                stand.append((b_sheet.cell(row=k, column=1).value, b_sheet.cell(row=k, column=j).value))


    wage = []
    """for row in cursor:
        stand.append(row)"""
        
    stand = standard(stand)
    
    for x, y in stand:
        y *= scale
        wage.append([x, y])

    return(wage)

# 社會因素
def callSociety(scale, prop):

    """query = "select neighborhood, " + prop +" from house.dbo.society;"
    cursor = conn.cursor()
    cursor.execute(query)"""
    stand = []
    wb = load_workbook('DSS資料v3.xlsx')
    a_sheet = wb.get_sheet_by_name('中壢')
    b_sheet = wb.get_sheet_by_name('桃園')
    for j in range(16, 18):
        if prop == a_sheet.cell(row=4, column=j).value:
            for i in range(6, 91):
                print(a_sheet.cell(row=i, column=j).value)
                stand.append((a_sheet.cell(row=i, column=1).value, a_sheet.cell(row=i, column=j).value))
            for k in range(6, 82):
                stand.append((b_sheet.cell(row=k, column=1).value, b_sheet.cell(row=k, column=j).value))

 
    society = []
    """for row in cursor:
        stand.append(row)"""
        
    stand = standard(stand)
    
    for x, y in stand:
        y *= scale
        society.append([x, y])

    return(society)

# 交通
def callTraffic(scale, prop):

    """query = "select neighborhood, " + prop +" from house.dbo.traffic;"
    cursor = conn.cursor()
    cursor.execute(query)"""
    stand = []
    wb = load_workbook('DSS資料v3.xlsx')
    a_sheet = wb.get_sheet_by_name('中壢')
    b_sheet = wb.get_sheet_by_name('桃園')
    for j in range(18, 23):
        if prop == a_sheet.cell(row=4, column=j).value:
            for i in range(6, 91):
                print(a_sheet.cell(row=i, column=j).value)
                stand.append((a_sheet.cell(row=i, column=1).value, a_sheet.cell(row=i, column=j).value))
            for k in range(6, 82):
                stand.append((b_sheet.cell(row=k, column=1).value, b_sheet.cell(row=k, column=j).value))

  
    traffic = []
    """for row in cursor:
        stand.append(row)"""
        
    stand = standard(stand)
    
    for x, y in stand:
        y *= scale
        traffic.append([x, y])

    return(traffic)

# 房屋資訊
def callHouse(scale, prop):

    """query = "select neighborhood, " + prop +" from house.dbo.house;"
    cursor = conn.cursor()
    cursor.execute(query)"""
    stand = []
    wb = load_workbook('DSS資料v3.xlsx')
    a_sheet = wb.get_sheet_by_name('中壢')
    b_sheet = wb.get_sheet_by_name('桃園')
    for j in range(23, 26):
        if prop == a_sheet.cell(row=4, column=j).value:
            for i in range(6, 91):
                print(a_sheet.cell(row=i, column=j).value)
                stand.append((a_sheet.cell(row=i, column=1).value, a_sheet.cell(row=i, column=j).value))
            for k in range(6, 82):
                stand.append((b_sheet.cell(row=k, column=1).value, b_sheet.cell(row=k, column=j).value))

    house = []
    """for row in cursor:
        stand.append(row)"""
        
    stand = standard(stand)
    
    for x, y in stand:
        y *= scale
        house.append([x, y])

    return(house)
print("1")
print(callLife(1,"park"))





